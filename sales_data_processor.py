#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
销售数据处理脚本 — Sales Data Processing Script
=================================================
功能说明：
  从 CSV 文件中读取销售数据，执行数据清洗（去重、补全缺失值、格式标准化），
  然后按日期和品类汇总统计，最后生成格式规范的 Excel 报告。

依赖库：
  - pandas    数据处理核心库
  - openpyxl  Excel 文件输出引擎

使用方法：
  python sales_data_processor.py <input.csv> [output.xlsx]

示例：
  python sales_data_processor.py sales_data.csv sales_report.xlsx
"""

import sys
import warnings
from datetime import datetime

import pandas as pd

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ============================================================================
# 配置区 — 可根据实际 CSV 列名调整
# ============================================================================

# 期望的列名映射（原始CSV列名 → 标准列名）
COLUMN_MAPPING = {
    "日期": "date",
    "Date": "date",
    "date": "date",
    "品类": "category",
    "Category": "category",
    "category": "category",
    "产品类别": "category",
    "销售额": "amount",
    "Amount": "amount",
    "amount": "amount",
    "销售金额": "amount",
    "金额": "amount",
    "数量": "quantity",
    "Quantity": "quantity",
    "quantity": "quantity",
    "销售数量": "quantity",
}

# 日期格式（自动尝试解析的格式列表）
DATE_FORMATS = [
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M:%S",
    "%m/%d/%Y",
    "%d/%m/%Y",
    "%Y年%m月%d日",
]


# ============================================================================
# 数据清洗函数
# ============================================================================

def load_and_clean_data(filepath: str) -> pd.DataFrame:
    """
    加载 CSV 并执行完整的数据清洗流程

    步骤：
      1. 读取原始 CSV
      2. 列名标准化（映射为标准英文列名）
      3. 删除完全重复的行
      4. 补全缺失值
      5. 格式标准化（日期、数值）
      6. 剔除无效数据
    """
    # ---------- 1. 读取 CSV ----------
    try:
        df = pd.read_csv(filepath, encoding="utf-8-sig")
    except UnicodeDecodeError:
        # 部分中文 CSV 使用 GBK 编码，fallback 尝试
        df = pd.read_csv(filepath, encoding="gbk")
    except FileNotFoundError:
        print(f"❌ 错误：文件未找到 — {filepath}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ 错误：读取 CSV 失败 — {e}")
        sys.exit(1)

    print(f"📄 原始数据：{len(df)} 行, {len(df.columns)} 列")
    print(f"   列名：{list(df.columns)}")

    # ---------- 2. 列名标准化 ----------
    df.rename(columns=COLUMN_MAPPING, inplace=True)
    # 仅保留我们关注的列（date, category, amount, quantity）
    keep_cols = [c for c in ["date", "category", "amount", "quantity"] if c in df.columns]
    if not keep_cols:
        print("❌ 错误：CSV 中未找到任何可识别的列名。请检查 COLUMN_MAPPING 配置。")
        sys.exit(1)
    df = df[keep_cols]

    print(f"📋 标准化列名后：{list(df.columns)}")

    # ---------- 3. 删除完全重复的行 ----------
    before = len(df)
    df.drop_duplicates(inplace=True)
    dup_removed = before - len(df)
    if dup_removed:
        print(f"🧹 去重：移除 {dup_removed} 条重复行")

    # ---------- 4. 补全缺失值 ----------
    missing_before = df.isnull().sum().to_dict()

    # 日期缺失 → 删除（无法补全）
    if "date" in df.columns:
        df.dropna(subset=["date"], inplace=True)

    # 品类缺失 → 标记为"未分类"
    if "category" in df.columns:
        df["category"].fillna("未分类", inplace=True)

    # 金额缺失 → 填充为 0
    if "amount" in df.columns:
        df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
        df["amount"].fillna(0.0, inplace=True)

    # 数量缺失 → 填充为 0
    if "quantity" in df.columns:
        df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce")
        df["quantity"].fillna(0, inplace=True)

    missing_after = df.isnull().sum().to_dict()
    for col in missing_before:
        if missing_before[col] > 0:
            print(f"🔧 缺失值处理 [{col}]：{missing_before[col]} → {missing_after.get(col, 0)}")

    # ---------- 5. 日期格式标准化 ----------
    if "date" in df.columns:
        parsed = None
        for fmt in DATE_FORMATS:
            try:
                parsed = pd.to_datetime(df["date"], format=fmt, errors="coerce")
                if parsed.notna().sum() > len(df) * 0.5:
                    break
            except Exception:
                continue
        if parsed is None or parsed.notna().sum() == 0:
            # 最终尝试 pandas 自动推断
            parsed = pd.to_datetime(df["date"], errors="coerce")

        invalid_dates = parsed.isna().sum()
        if invalid_dates:
            print(f"⚠️  日期解析失败 {invalid_dates} 行（已删除）")
        df["date"] = parsed
        df.dropna(subset=["date"], inplace=True)

    # ---------- 6. 剔除金额/数量为负的异常数据 ----------
    if "amount" in df.columns:
        neg = (df["amount"] < 0).sum()
        if neg:
            print(f"⚠️  剔除负金额数据 {neg} 行")
            df = df[df["amount"] >= 0]

    if "quantity" in df.columns:
        neg_q = (df["quantity"] < 0).sum()
        if neg_q:
            print(f"⚠️  剔除负数量数据 {neg_q} 行")
            df = df[df["quantity"] >= 0]

    # 重置索引
    df.reset_index(drop=True, inplace=True)

    print(f"✅ 清洗完成：{len(df)} 行有效数据")
    return df


# ============================================================================
# 汇总统计函数
# ============================================================================

def generate_summary(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """
    按日期和品类生成多维度汇总统计表

    返回：
      {
        "by_date":      按日期汇总,
        "by_category":  按品类汇总,
        "by_date_cat":  按日期+品类交叉汇总,
        "overview":     总览统计,
      }
    """
    summaries = {}

    # ---------- 按日期汇总 ----------
    if "date" in df.columns:
        by_date = df.groupby(df["date"].dt.date).agg(
            订单数=("amount", "count"),
            销售总额=("amount", "sum"),
            平均单额=("amount", "mean"),
            总数量=("quantity", "sum") if "quantity" in df.columns else (),
        )
        # 如果 quantity 存在，保留完整 agg
        if "quantity" in df.columns:
            by_date = df.groupby(df["date"].dt.date).agg(
                订单数=("amount", "count"),
                销售总额=("amount", "sum"),
                平均单额=("amount", "mean"),
                总数量=("quantity", "sum"),
            )
        by_date = by_date.reset_index()
        by_date.columns = ["日期", "订单数", "销售总额", "平均单额", "总数量"]
        by_date["销售总额"] = by_date["销售总额"].round(2)
        by_date["平均单额"] = by_date["平均单额"].round(2)
        summaries["by_date"] = by_date

    # ---------- 按品类汇总 ----------
    if "category" in df.columns:
        by_cat = df.groupby("category").agg(
            订单数=("amount", "count"),
            销售总额=("amount", "sum"),
            平均单额=("amount", "mean"),
            总数量=("quantity", "sum") if "quantity" in df.columns else (),
        )
        if "quantity" in df.columns:
            by_cat = df.groupby("category").agg(
                订单数=("amount", "count"),
                销售总额=("amount", "sum"),
                平均单额=("amount", "mean"),
                总数量=("quantity", "sum"),
            )
        by_cat = by_cat.reset_index()
        by_cat.columns = ["品类", "订单数", "销售总额", "平均单额", "总数量"]
        by_cat["销售总额"] = by_cat["销售总额"].round(2)
        by_cat["平均单额"] = by_cat["平均单额"].round(2)
        summaries["by_category"] = by_cat

    # ---------- 按日期+品类交叉汇总 ----------
    if "date" in df.columns and "category" in df.columns:
        pivot = df.pivot_table(
            values="amount",
            index=df["date"].dt.date,
            columns="category",
            aggfunc="sum",
            fill_value=0,
        ).round(2)
        pivot.index.name = "日期"
        pivot.columns.name = None
        # 添加合计行
        pivot.loc["合计"] = pivot.sum()
        summaries["by_date_category"] = pivot

    # ---------- 总览统计 ----------
    overview = pd.DataFrame({
        "指标": [
            "数据范围",
            "总订单数",
            "总销售额",
            "平均单额",
            "品类数量",
            "天数",
        ],
        "数值": [
            f"{df['date'].min().date()} ~ {df['date'].max().date()}"
            if "date" in df.columns else "N/A",
            len(df),
            f"{df['amount'].sum():.2f}" if "amount" in df.columns else "N/A",
            f"{df['amount'].mean():.2f}" if "amount" in df.columns else "N/A",
            df["category"].nunique() if "category" in df.columns else "N/A",
            df["date"].dt.date.nunique() if "date" in df.columns else "N/A",
        ],
    })
    summaries["overview"] = overview

    return summaries


# ============================================================================
# Excel 报告输出
# ============================================================================

def write_excel_report(
    summaries: dict[str, pd.DataFrame],
    output_path: str,
    raw_data: pd.DataFrame = None,
):
    """
    将汇总统计写入格式规范的 Excel 文件

    使用 openpyxl 引擎进行单元格格式美化：
      - 表头加粗、浅蓝色填充
      - 数值列自动列宽
      - 冻结首行
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ---------- Sheet 1: 总览 ----------
        if "overview" in summaries:
            summaries["overview"].to_excel(writer, sheet_name="总览", index=False)

        # ---------- Sheet 2: 按日期汇总 ----------
        if "by_date" in summaries:
            summaries["by_date"].to_excel(writer, sheet_name="按日期汇总", index=False)

        # ---------- Sheet 3: 按品类汇总 ----------
        if "by_category" in summaries:
            summaries["by_category"].to_excel(writer, sheet_name="按品类汇总", index=False)

        # ---------- Sheet 4: 交叉汇总 ----------
        if "by_date_category" in summaries:
            summaries["by_date_category"].to_excel(writer, sheet_name="日期×品类交叉汇总")

        # ---------- Sheet 5: 清洗后明细 ----------
        if raw_data is not None:
            raw_data.to_excel(writer, sheet_name="清洗后明细", index=False)

    # ========== 格式美化（使用 openpyxl） ==========
    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 交替行填充（斑马纹）
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for ws in wb.worksheets:
        # --- 表头格式 ---
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # --- 数据行格式 ---
        for row_idx in range(2, ws.max_row + 1):
            fill = even_fill if row_idx % 2 == 0 else odd_fill
            for cell in ws[row_idx]:
                cell.alignment = cell_alignment
                cell.border = thin_border
                cell.fill = fill
                # 数值列保留两位小数
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"

        # --- 自动列宽 ---
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    val = str(cell.value) if cell.value is not None else ""
                    # 中文字符按 2 个宽度计算
                    cjk_count = sum(1 for c in val if "\u4e00" <= c <= "\u9fff")
                    length = len(val) + cjk_count
                    if length > max_len:
                        max_len = length
            adjusted_width = min(max_len + 4, 40)
            ws.column_dimensions[col_letter].width = max(adjusted_width, 10)

        # --- 冻结首行 ---
        ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"📁 Excel 报告已保存：{output_path}")


# ============================================================================
# 主流程
# ============================================================================

def main():
    if len(sys.argv) < 2:
        print("用法：python sales_data_processor.py <input.csv> [output.xlsx]")
        print("示例：python sales_data_processor.py sales_data.csv sales_report.xlsx")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else "sales_report.xlsx"

    print("=" * 60)
    print("🚀 销售数据处理脚本 — 开始运行")
    print(f"   输入：{input_file}")
    print(f"   输出：{output_file}")
    print("=" * 60)

    # 1. 加载并清洗数据
    df = load_and_clean_data(input_file)

    # 2. 生成汇总统计
    summaries = generate_summary(df)

    # 3. 输出 Excel 报告
    write_excel_report(summaries, output_file, raw_data=df)

    # 4. 打印终端摘要
    print("\n" + "=" * 60)
    print("📊 报表摘要")
    print("=" * 60)
    if "by_category" in summaries:
        print("\n按品类销售汇总：")
        print(summaries["by_category"].to_string(index=False))
    if "by_date" in summaries:
        print("\n按日期销售汇总：")
        print(summaries["by_date"].to_string(index=False))
    print("\n✅ 全部完成！")


if __name__ == "__main__":
    main()
